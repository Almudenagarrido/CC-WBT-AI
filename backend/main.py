import os
import uuid
import stat
import copy
import json
import shutil
import tempfile
import openpyxl
import traceback
import pandas as pd
from shutil import copyfile
from pydantic import BaseModel
from typing import List, Dict, Any
from excel_formula_engine import ExcelFormulaProcessor
from fastapi.responses import FileResponse, JSONResponse
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
JSON_FORMULAS = "formulas_map.json"

def load_config():
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)


app = FastAPI()
excel_processor = ExcelFormulaProcessor()


# Countries (GET, POST, COPY, DELETE)
class CountryRequest(BaseModel):
    name: str
    tax_rate: int
    inflation: int

@app.get("/countries")
def get_countries():
    config = load_config()
    return {"countries": config["COUNTRIES"]}

@app.post("/countries")
def add_country(request: CountryRequest):
    config = load_config()
    new_country = request.name.strip()
    tax_rate = request.tax_rate
    inflation = request.inflation

    if not new_country:
        raise HTTPException(status_code=400, detail="Country name empty")
    if new_country in config["COUNTRIES"]:
        raise HTTPException(status_code=400, detail="Country already exists")

    config["COUNTRIES"].append(new_country)
    config["TAX_RATES"][new_country] = tax_rate
    config["INFLATIONS"][new_country] = inflation
    save_config(config)
    return {"message": f"Country '{new_country}' added."}

@app.post("/countries/{country}")
def create_templates_for_country(country: str):
    dst = os.path.join(BASE_DIR, country)

    if os.path.exists(dst):
        return {"message": f"Templates for '{country}' already exist."}
    
    try:
        templates_path = os.path.join(BASE_DIR, "{templates}")
        shutil.copytree(templates_path, dst)
        return {"message": f"Templates for '{country}' created."}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error copying templates: {e}")

def remove_readonly(func, path, excinfo):
    os.chmod(path, stat.S_IWRITE)
    func(path)

@app.get("/download-country")
def download_country_files(country: str, background_tasks: BackgroundTasks):
    temp_dir = tempfile.mkdtemp()
    try:
        folder_path = os.path.join(BASE_DIR, country)        
        internal_folder = os.path.join(temp_dir, country)
        os.makedirs(internal_folder, exist_ok=True)

        files_to_include = [
            f for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f)) and "{model}" not in f and "{template}" not in f
        ]

        for filename in files_to_include:
            src = os.path.join(folder_path, filename)
            dst = os.path.join(internal_folder, filename)
            shutil.copy(src, dst)

        zip_base = os.path.join(tempfile.gettempdir(), f"{country}_files")
        zip_path = shutil.make_archive(zip_base, 'zip', internal_folder)

        background_tasks.add_task(os.remove, zip_path)
        shutil.rmtree(temp_dir)

        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=os.path.basename(zip_path)
        )

    except Exception as e:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/countries")
def delete_country(request: CountryRequest):
    config = load_config()
    country = request.name.strip()
    
    if country not in config["COUNTRIES"]:
        raise HTTPException(status_code=404, detail="Country not found")

    config["COUNTRIES"].remove(country)
    for section in ["COUNTRY_YEAR_RANGES", "TAX_RATES", "INFLATIONS","FUELS", "MODELS", "CONSUMER_TYPES"]:
        if country in config.get(section, {}):
            del config[section][country]
    save_config(config)

    folder = os.path.join(BASE_DIR, country)
    if os.path.exists(folder) and os.path.isdir(folder):
        try:
            shutil.rmtree(folder, onerror=remove_readonly)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error deleting folder: {e}")
        
    return {"message": f"Country '{country}' deleted."}


# Fuels (GET, POST, DELETE)
class FuelRequest(BaseModel):
    fuel: str
    country: str

def sort_fuels(fuels):
    config = load_config()
    return sorted(fuels, key=lambda x: (x not in config["ELECTRICITY_VARIANTS"], x))

@app.get("/fuels")
def get_fuels(key, country):
    config = load_config()
    key = key.strip()
    country = country.strip()

    if country not in config["FUELS"]:

        if "template" not in config["FUELS"]:
            raise HTTPException(status_code=500, detail="Template country not available.")
        
        config["FUELS"][country] = copy.deepcopy(config["FUELS"]["template"])

    if key not in config["FUELS"][country]:
        raise HTTPException(status_code=404, detail=f"Key '{key}' not found in country '{country}'.")
    
    save_config(config)
    return {"fuels": sort_fuels(config["FUELS"][country][key])}

@app.post("/add-fuels")
def add_fuel(request: FuelRequest):
    config = load_config()
    fuel = request.fuel.strip()
    country = request.country.strip()

    if not fuel:
        raise HTTPException(status_code=400, detail="Fuel name is empty")

    if country not in config["FUELS"]:
        raise HTTPException(status_code=404, detail=f"Country '{country}' not found in fuels.")

    added_to = []
    for key in config["FUELS"][country]:
        if fuel == "Electricity":
            if key == "carbon":
                continue

            if key == "expanded":
                values = ["Electricity & E-Cooking", "Electricity (Just access)"]
            elif key == "more_expanded":
                values = ["Electricity (Only E-Cooking)", "Electricity & E-Cooking", "Electricity (Just access)"]
            else:
                values = ["Electricity"]
        else:
            values = [fuel]

        for v in values:
            if v not in config["FUELS"][country][key]:
                config["FUELS"][country][key].append(v)
                added_to.append((key, v))

    if not added_to:
        return {"message": f"Fuel already present in all lists for {country}."}
    
    save_config(config)
    return {"message": f"Fuel added to {country}: {added_to}"}

@app.delete("/delete-fuels")
def delete_fuel(request: FuelRequest):
    config = load_config()
    fuel = request.fuel.strip()
    country = request.country.strip()

    if not fuel:
        raise HTTPException(status_code=400, detail="Fuel keyword is empty")
    
    if country not in config["FUELS"]:
        raise HTTPException(status_code=404, detail=f"Country '{country}' not found in fuels.")

    if fuel.lower().startswith("electricity"):
        to_delete = config["ELECTRICITY_VARIANTS"]
    else:
        to_delete = {fuel}

    deleted = []
    for key in config["FUELS"][country]:
        original = config["FUELS"][country][key]
        filtered = [f for f in original if f not in to_delete]
        removed = set(original) - set(filtered)
        config["FUELS"][country][key] = filtered
        deleted.extend(removed)

    if not deleted:
        raise HTTPException(status_code=404, detail="No matching fuel entries found")

    save_config(config)
    return {"message": f"Removed entries: {sorted(set(deleted))}"}


# Models (GET, POST, DOWNLOAD, DELETE)
class ModelRequest(BaseModel):
    country: str
    model: str

@app.get("/models")
def get_models(country):
    config = load_config()
    country = country.strip()

    if country not in config["MODELS"]:
        config["MODELS"][country] = []
    
    return {"models": config["MODELS"][country]}

@app.post("/model")
async def create_model(country: str, model: str, start_year: int, end_year: int):
    config = load_config()
    country = country.strip()
    model = model.strip()

    if model.lower() == "bau":
        if country not in config["COUNTRY_YEAR_RANGES"]:
            config["COUNTRY_YEAR_RANGES"][country] = {
                "start": start_year,
                "end": end_year
            }
            config["MODELS"][country] = ["BAU"]
    
    else:
        expected_range = config["COUNTRY_YEAR_RANGES"][country]
        if start_year != expected_range["start"] or end_year != expected_range["end"]:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Year range mismatch. Expected "
                    f"{expected_range['start']}–{expected_range['end']}, "
                    f"got {start_year}–{end_year}."
                )
            )
        
        for route in config["ROUTES"]:
            src_path = os.path.join(BASE_DIR, country, route)
            dst_path = os.path.join(BASE_DIR, country, route.format(model=model))

            if os.path.exists(src_path):
                shutil.copy(src_path, dst_path)
            else:
                raise HTTPException(status_code=404, detail=f"Template file not found: {src_path}")
                
        config["MODELS"][country].append(model)

    save_config(config)
    return {"message": f"Model '{model}' created successfully for country '{country}'."}

@app.get("/download-model")
def download_model_files(country:str, model: str, background_tasks: BackgroundTasks):
    config = load_config()
    temp_dir = tempfile.mkdtemp()

    try:
        folder_path = os.path.join(BASE_DIR, country)
        files_to_copy = []

        if model.lower() != "bau":
            for route in config["ROUTES"]:
                filename = route.format(model=model)
                full_path = os.path.join(folder_path, filename)
                if not os.path.exists(full_path):
                    shutil.rmtree(temp_dir)
                files_to_copy.append(full_path)

        for shared_route in config["SHARED_ROUTES"]:
            full_path = os.path.join(folder_path, shared_route)
            if os.path.exists(full_path):
                files_to_copy.append(full_path)

        for path in files_to_copy:
            shutil.copy(path, os.path.join(temp_dir, os.path.basename(path)))

        zip_base = os.path.join(tempfile.gettempdir(), f"{country}_{model}_files")
        zip_path = shutil.make_archive(zip_base, 'zip', temp_dir)

        background_tasks.add_task(os.remove, zip_path)
        shutil.rmtree(temp_dir)
        
        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=os.path.basename(zip_path)
        )

    except Exception as e:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        raise HTTPException(status_code=500, detail=str(e))

def get_year_columns(sheet):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str) and val.strip().lower() == "baseline":
                year_columns = {}
                year_columns[col] = "baseline"
                for c in range(col + 1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=row, column=c).value
                    try:
                        year = int(cell_val)
                        year_columns[c] = year
                    except (TypeError, ValueError):
                        continue
                return row, year_columns
    return None, {}

def drop_out_of_range_years_from_workbook(wb, sheet_name, start_year, end_year):
    ws = wb[sheet_name]
    row_idx, year_columns = get_year_columns(ws)

    if not year_columns or row_idx is None:
        return

    cols_to_delete = []
    baseline_year = start_year
    
    for col_idx, year in year_columns.items():
        if year == "baseline":
            continue
        else:
            if not isinstance(year, int):
                continue
            if year < start_year or year > end_year:
                cols_to_delete.append(col_idx)
            elif year == baseline_year:
                cols_to_delete.append(col_idx)

    for col in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col)

def sync_sheets_with_fuels(country, route, template_route, expected_sheets):
    config = load_config()
    full_path = os.path.join(BASE_DIR, route)
    template_path = os.path.join(BASE_DIR, template_route)

    if not os.path.isfile(full_path):
        if not os.path.isfile(template_path):
            raise FileNotFoundError("Template not found")
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        copyfile(template_path, full_path)
    
    wb = openpyxl.load_workbook(full_path)
    wb_template = openpyxl.load_workbook(template_path)
    changed = False
    
    for fuel_sheet in expected_sheets:
        if fuel_sheet not in wb.sheetnames:
            template_source_sheet = None

            if fuel_sheet in wb_template.sheetnames:
                template_source_sheet = wb_template[fuel_sheet]
            elif "LPG" in wb_template.sheetnames:
                template_source_sheet = wb_template["LPG"]
            elif expected_sheets and expected_sheets[0] in wb_template.sheetnames:
                template_source_sheet = wb_template[expected_sheets[0]]
            elif wb_template.sheetnames:
                template_source_sheet = wb_template[wb_template.sheetnames[0]]
            else:
                raise ValueError("No sheets available in template file")
            
            new_sheet = wb.create_sheet(fuel_sheet)
            for row in template_source_sheet.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value
                    if cell.has_style:
                        new_sheet[cell.coordinate]._style = cell._style

            for col_letter, dimension in template_source_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = dimension.width
                new_sheet.column_dimensions[col_letter].hidden = dimension.hidden
            
            changed = True

    year_range = config["COUNTRY_YEAR_RANGES"].get(country)    
    if year_range:
        start_year, end_year = year_range["start"], year_range["end"]
        for sheet_name in wb.sheetnames:
            drop_out_of_range_years_from_workbook(wb, sheet_name, start_year, end_year)
        changed = True

    sheets_to_delete = [sheet for sheet in wb.sheetnames if sheet not in expected_sheets]
    
    for sheet in sheets_to_delete:
        std = wb[sheet]
        wb.remove(std)
        changed = True

    if changed:
        wb.save(full_path)

    wb.close()
    wb_template.close()

    return changed

def add_carbon_credits_sheet(country, download_path, template_path, model):
    try:
        carbon_sheet = "Carbon Credits"
        config = load_config()
        
        wb_download = openpyxl.load_workbook(download_path)
        wb_template = openpyxl.load_workbook(template_path)
        
        if carbon_sheet in wb_template.sheetnames and carbon_sheet not in wb_download.sheetnames:
            template_sheet = wb_template[carbon_sheet]
            new_sheet = wb_download.create_sheet(carbon_sheet)
            
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value
                    if cell.has_style:
                        new_sheet[cell.coordinate]._style = cell._style
        
        if model:
            for sheet_name in wb_download.sheetnames:
                worksheet = wb_download[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and "{model}" in cell.value:
                            cell.value = cell.value.replace("{model}", model)

        year_range = config["COUNTRY_YEAR_RANGES"].get(country)
        if year_range:
            start_year, end_year = year_range["start"], year_range["end"]
            for sheet_name in wb_download.sheetnames:
                drop_out_of_range_years_from_workbook(wb_download, sheet_name, start_year, end_year)
            
        wb_download.save(download_path)
        
        wb_download.close()
        wb_template.close()
        
    except Exception as e:
        try:
            wb_download.close()
            wb_template.close()
        except:
            pass
        raise e
  
@app.get("/download-template")
def download_template_file(country, template, model, key_fuels):
    config = load_config()
    expected_sheets = config["FUELS"].get(country, {}).get(key_fuels, [])
    temp_file_path = None
    
    try:
        template_path = os.path.join(BASE_DIR, country, template)
        
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Template file not found")
        
        temp_dir = tempfile.gettempdir()
        temp_filename = f"synced_{uuid.uuid4().hex[:8]}_{template}"
        temp_file_path = os.path.join(temp_dir, temp_filename)
        
        sync_sheets_with_fuels(
            country=country,
            route=temp_file_path,
            template_route=template_path,
            expected_sheets=expected_sheets
        )
        add_carbon_credits_sheet(country, temp_file_path, template_path, model)

        return FileResponse(
            temp_file_path,
            media_type="application/vnd.ms-excel",
            filename=template
        )
        
    except Exception as e:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass
        raise HTTPException(status_code=500, detail=str(e))
 
@app.post("/upload-template")
async def upload_template_file(country: str, model: str, file: UploadFile = File(...)):
    try:
        expected_name = f"upload-BAU.xlsx" if model.lower() == "bau" else f"upload-{model}.xlsx"
        
        if file.filename != expected_name:
            raise HTTPException(
                status_code=400
            )
        
        country_dir = os.path.join(BASE_DIR, country)
        os.makedirs(country_dir, exist_ok=True)
        file_path = os.path.join(country_dir, file.filename)
        
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        return {"status": "success", "filename": file.filename}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.delete("/model")
def delete_model(request: ModelRequest):
    config = load_config()
    country = request.country.strip()
    model = request.model.strip()

    if model.lower() == "bau":
        all_models = config["MODELS"].get(country, [])

        if country in config.get("CONSUMER_TYPES", {}):
            config["CONSUMER_TYPES"].pop(country, None)

        for m in all_models:
            for route in config["ROUTES"]:
                file_path = os.path.join(BASE_DIR, country, route.format(model=m))
                if os.path.exists(file_path):
                    os.remove(file_path)
            
            upload_path = os.path.join(BASE_DIR, country, f"upload-{m}.xlsx")
            if os.path.exists(upload_path):
                os.remove(upload_path)

        destination_upload_path = os.path.join(BASE_DIR, country, "upload-BAU.xlsx")
        source_upload_path = os.path.join(BASE_DIR, "{templates}", "upload-BAU.xlsx")
        shutil.copy2(source_upload_path, destination_upload_path)

        config["MODELS"].pop(country, None)
        config["COUNTRY_YEAR_RANGES"].pop(country, None)
    else:
        for route in config["ROUTES"]:
            file_path = os.path.join(BASE_DIR, country, route.format(model=model))
            if os.path.exists(file_path):
                os.remove(file_path)
        
        upload_path = os.path.join(BASE_DIR, country, f"upload-{model}.xlsx")
        if os.path.exists(upload_path):
            os.remove(upload_path)

        config["MODELS"][country].remove(model)
        if country in config.get("CONSUMER_TYPES", {}):
            if model in config["CONSUMER_TYPES"][country]:
                config["CONSUMER_TYPES"][country].pop(model, None)

    save_config(config)
    return {"message": f"{model}' deleted successfully for country '{country}'."}
    
    
# Consumers (GET, POST, DELETE)
class ConsumerRequest(BaseModel):
    country: str
    model: str
    consumer: str

@app.get("/consumers")
def get_consumers(country, model):
    config = load_config()
    country = country.strip()
    model = model.strip()

    if country not in config["CONSUMER_TYPES"]:

        if "template" not in config["CONSUMER_TYPES"]:
            raise HTTPException(status_code=500, detail="Template country not available.")
        
        config["CONSUMER_TYPES"][country] = copy.deepcopy(config["CONSUMER_TYPES"]["template"])

    if model not in config["CONSUMER_TYPES"][country]:

        if "model" not in config["CONSUMER_TYPES"][country]:
            raise HTTPException(status_code=500, detail="Template country not available.")
        
        config["CONSUMER_TYPES"][country][model] = copy.deepcopy(config["CONSUMER_TYPES"][country]["model"])
    
    save_config(config)
    return {"consumers": config["CONSUMER_TYPES"][country][model]}

@app.post("/add-consumer")
def add_consumer(request: ConsumerRequest):
    config = load_config()
    country = request.country.strip()
    model = request.model.strip()
    consumer = request.consumer.strip()
    
    if not consumer:
        raise HTTPException(status_code=400, detail="Consumer type name is empty")
    
    if "CONSUMER_TYPES" not in config:
        config["CONSUMER_TYPES"] = {}
    if country not in config["CONSUMER_TYPES"]:
        config["CONSUMER_TYPES"][country] = {}
    if model not in config["CONSUMER_TYPES"][country]:
        config["CONSUMER_TYPES"][country][model] = []

    if consumer not in config["CONSUMER_TYPES"][country][model]:
        config["CONSUMER_TYPES"][country][model].append(consumer)
        save_config(config)
    
    return {"message": f"Consumer type '{consumer}' added to model '{model}' for country '{country}'"}

@app.delete("/delete-consumer")
def delete_consumer(request: ConsumerRequest):
    config = load_config()
    country = request.country.strip()
    model = request.model.strip()
    consumer = request.consumer.strip()

    if not consumer:
        raise HTTPException(status_code=400, detail="Consumer type name is empty")
    
    if country not in config["CONSUMER_TYPES"]:
        raise HTTPException(status_code=404, detail=f"Country '{country}' not found in consumer types.")
    
    if model not in config["CONSUMER_TYPES"][country]:
        raise HTTPException(status_code=404, detail=f"Model '{model}' not found in consumer types for country '{country}'.")

    if consumer in config["CONSUMER_TYPES"][country][model]:
        config["CONSUMER_TYPES"][country][model].remove(consumer)
        save_config(config)

    return {"message": f"Removed entry for consumer '{consumer}' for '{model}' in country '{country}'."}


# Excel Sheet (REMOVE, EXPAND, GET, POST, RESET)
class SheetUpdate(BaseModel):
    country: str
    route: str
    template_route: str
    sheet_name: str
    data: List[Dict[str, Any]]
    models: List[str]
    key_fuels: str

@app.post("/remove-models")
async def remove_models(remove_request: SheetUpdate):
    try:
        route = remove_request.route
        sheet_name = remove_request.sheet_name
        models = remove_request.models
        full_path = os.path.join(BASE_DIR, route)
        
        wb = openpyxl.load_workbook(full_path)            
        ws = wb[sheet_name]
        
        rows_to_delete = []
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and isinstance(cell_val, str):
                for model in models:
                    if model in cell_val and "{model}" not in cell_val:
                        rows_to_delete.append(row_idx)
                        break
        
        for row_idx in sorted(set(rows_to_delete), reverse=True):
            ws.delete_rows(row_idx)
        
        wb.save(full_path)
        wb.close()

        return {
            "message": f"Removed {len(rows_to_delete)} rows for {len(models)} models", 
            "removed": True,
        }
            
    except Exception as e:
        if 'wb' in locals():
            wb.close()
        raise HTTPException(status_code=500, detail=f"Error removing models: {str(e)}")

@app.post("/expand-sheet")
async def expand_sheet(expand_request: SheetUpdate):
    try:
        route = expand_request.route
        sheet_name = expand_request.sheet_name
        models = expand_request.models

        full_path = os.path.join(BASE_DIR, route)
        wb = openpyxl.load_workbook(full_path)            
        ws = wb[sheet_name]
        
        template_rows = []
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and isinstance(cell_val, str) and "{model}" in cell_val:
                row_data = {
                    'original_idx': row_idx,
                    'cells': [ws.cell(row=row_idx, column=col_idx) 
                             for col_idx in range(1, ws.max_column + 1)],
                    'content_type': cell_val.split('{model}')[0].strip()
                }
                template_rows.append(row_data)
        
        if not template_rows:
            wb.close()
            return {"expanded": False}
        
        template_rows.sort(key=lambda x: x['original_idx'])
        current_offset = 0
        
        for template in template_rows:
            original_idx = template['original_idx'] + current_offset
            cells_to_copy = template['cells']
                        
            for model in models:
                exists = False
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                    first_val = row[0] if row else None
                    if first_val and isinstance(first_val, str) and model in first_val:
                        if template['content_type'] in first_val:
                            exists = True
                            break
                
                if exists:
                    continue
                
                insert_pos = original_idx + 1
                ws.insert_rows(insert_pos)
                current_offset += 1
                
                for col_idx, orig_cell in enumerate(cells_to_copy, start=1):
                    new_cell = ws.cell(row=insert_pos, column=col_idx)
                    
                    if (orig_cell.value and isinstance(orig_cell.value, str) and 
                        "{model}" in orig_cell.value):
                        new_cell.value = orig_cell.value.replace("{model}", model)
                    else:
                        new_cell.value = orig_cell.value
                    
                    if orig_cell.has_style:
                        new_cell._style = orig_cell._style

        wb.save(full_path)
        wb.close()
        return {"expanded": True}
            
    except Exception as e:
        traceback.print_exc()
        if 'wb' in locals():
            wb.close()
        raise e

@app.get("/get-sheet")
async def get_sheet(country, route, template_route, sheet_name, key_fuels):

    config = load_config()
    fuels = config["FUELS"].get(country, {}).get("normal", [])
    expected_sheets = config["FUELS"].get(country, {}).get(key_fuels, [])
    models = config["MODELS"].get(country, [])

    if sheet_name not in expected_sheets:
        raise HTTPException(status_code=400, detail="Sheet name not allowed for this fuel and country")

    is_carbon_credits = "carbon" in route.lower() or "carbon" in template_route.lower()
    
    full_path = os.path.join(BASE_DIR, route)
    template_full_path = os.path.join(BASE_DIR, template_route)

    if not os.path.isfile(full_path):
        if not os.path.isfile(template_full_path):
            raise HTTPException(status_code=404, detail="Template file not found")
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        copyfile(template_full_path, full_path)

    wb = openpyxl.load_workbook(full_path)

    if is_carbon_credits:
        year_range = config["COUNTRY_YEAR_RANGES"].get(country)    
        if year_range:
            start_year, end_year = year_range["start"], year_range["end"]
            drop_out_of_range_years_from_workbook(wb, sheet_name, start_year, end_year)
            wb.save(full_path)
    else:
        try:
            sync_sheets_with_fuels(country, route, template_route, expected_sheets)
        except Exception as e:
            wb.close()
            raise HTTPException(status_code=500, detail=f"Error syncing sheets: {str(e)}")

    wb.close()
    excel_processor.clear_workbook_cache()
    excel_processor.apply_formulas(
        file_path=route,
        formulas_json_path=JSON_FORMULAS,
        country=country,
        models=models,
        fuels=fuels,
        expected_sheets=expected_sheets
    )

    df = pd.read_excel(full_path, sheet_name=sheet_name, engine="openpyxl")
    df = df.where(pd.notnull(df), None)

    return JSONResponse({"sheet": df.to_dict(orient="records")})

@app.post("/save-sheet")
async def save_sheet(update: SheetUpdate):
    excel_path = os.path.join(BASE_DIR, update.route)
    
    if not os.path.isfile(excel_path):
        raise HTTPException(status_code=404, detail="File not found")

    wb = None
    original_sheet = None
    temp_sheet = None
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        if update.sheet_name in wb.sheetnames:
            original_sheet = wb[update.sheet_name]
            
            temp_sheet = wb.create_sheet("temp_style_sheet")
            for row in original_sheet.iter_rows():
                for cell in row:
                    temp_cell = temp_sheet[cell.coordinate]
                    temp_cell.value = cell.value
                    if cell.has_style:
                        temp_cell._style = cell._style
            
            wb.remove(original_sheet)
        
        new_sheet = wb.create_sheet(update.sheet_name)
        df_data = pd.DataFrame(update.data)
        
        if original_sheet:
            for cell in original_sheet[1]:
                new_cell = new_sheet.cell(row=1, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style

        for r_idx, row in enumerate(df_data.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
                
                if temp_sheet:
                    temp_cell = temp_sheet.cell(row=r_idx, column=c_idx)
                    if temp_cell.has_style:
                        cell._style = temp_cell._style
        
        if original_sheet:
            for col_letter, dimension in original_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = dimension.width
                new_sheet.column_dimensions[col_letter].hidden = dimension.hidden
        
        if temp_sheet:
            wb.remove(temp_sheet)
        wb.save(excel_path)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving sheet: {str(e)}")
    finally:
        if wb:
            wb.close()
    
    return {"message": f"Sheet '{update.sheet_name}' updated with preserved styles."}

@app.post("/reset-sheet")
async def reset_sheet(update: SheetUpdate):
    excel_path = os.path.join(BASE_DIR, update.route)
    template_path = os.path.join(BASE_DIR, update.template_route)

    if not os.path.isfile(template_path):
        raise HTTPException(status_code=404, detail="Template file not found")

    if not os.path.isfile(excel_path):
        raise HTTPException(status_code=404, detail="Target file not found")

    try:
        wb_target = openpyxl.load_workbook(excel_path)
        wb_template = openpyxl.load_workbook(template_path)
        
        template_source_sheet = None
        if update.sheet_name in wb_template.sheetnames:
            template_source_sheet = wb_template[update.sheet_name]
        elif "LPG" in wb_template.sheetnames:
            template_source_sheet = wb_template["LPG"]
        elif wb_template.sheetnames:
            template_source_sheet = wb_template[wb_template.sheetnames[0]]

        if update.sheet_name in wb_target.sheetnames:
            target_sheet = wb_target[update.sheet_name]
            wb_target.remove(target_sheet)

        new_sheet = wb_target.create_sheet(update.sheet_name)
        
        for row in template_source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = cell._style
        
        wb_target.save(excel_path)
        wb_target.close()
        wb_template.close()
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error resetting sheet: {str(e)}")
    
    return {"message": f"Sheet '{update.sheet_name}' in '{update.route}' reset successfully to template."}