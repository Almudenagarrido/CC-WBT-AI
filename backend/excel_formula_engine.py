import os
import json
import traceback
from copy import deepcopy
from itertools import product
from functools import lru_cache
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
OPS_MAP = {
    "addition": lambda *args: sum(args),
    "subtraction": lambda x, y: x - y,
    "multiply": lambda x, y: x * y,
    "multiply_per": lambda x, y: (x * y)/100,
    "divide": lambda x, y: x / y,
    "safe_divide": lambda x, y: (x / y) if y != 0 else 0,
    "safe_divide_minus": lambda x, y, z: (x / y - z) if y != 0 else 0,
    "gt": lambda x, y: x > y,
    "gt_eq": lambda x, y: x >= y,
    "lt": lambda x, y: x < y,
    "lt_eq": lambda x, y: x <= y,
    "equal": lambda x, y: x == y,
    "if": lambda condition, true_val, false_val: true_val if condition else false_val,
    "copy": lambda x: x,
    "min": lambda x, y: min(x, y),
    "max": lambda x, y: max(x, y),
    "abs": lambda x: abs(x),
    "negative": lambda x: - x,
    "percentage": lambda x: x*100,
    "int": lambda x: int(x)
}


class ExcelFormulaProcessor:
    
    def __init__(self):
        self.previously_calculated = {}
        self.amount_targets = ["How much to be financed", "Equity", "Equity - GRID", "Equity - OFF-GRID", "Grants", "Grants - GRID", "Grants - OFF-GRID","Debt", "Debt - GRID", "Debt - OFF-GRID", "WACC", "WACC - GRID", "WACC - OFF-GRID"]
    
    @lru_cache(maxsize=10)
    def _get_workbook(self, file_path, read_only=False, data_only=False):
        return load_workbook(file_path, read_only=read_only, data_only=data_only)
    
    def clear_workbook_cache(self):
        self._get_workbook.cache_clear()

    def _is_numeric_value(self, value):
        
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value.replace(",", ".").strip())
                return True
            except ValueError:
                return False
        return False

    def _convert_to_numeric(self, value):
        
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", ".").strip())
            except ValueError:
                return 0
        return 0
    
    def _apply_number_formatting(self, ws, year_range):

        start_col = None
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, int) and year_range['start'] <= cell.value <= year_range['end']:
                    consecutive = True
                    for offset in range(1, 5):
                        next_col_idx = cell.column + offset - 1
                        if next_col_idx >= ws.max_column:
                            break
                        next_cell_value = ws.cell(row=cell.row, column=next_col_idx + 1).value
                        expected_year = cell.value + offset
                        if not (isinstance(next_cell_value, int) and next_cell_value == expected_year):
                            consecutive = False
                            break
                    if consecutive:
                        start_col = cell.column
                        break
            if start_col:
                break

        if not start_col:
            return

        year_header_rows = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_value = row[start_col - 1].value
            if isinstance(cell_value, int) and year_range['start'] <= cell_value <= year_range['end']:
                year_header_rows.add(row[0].row)
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if row[0].row in year_header_rows:
                continue
            for col_idx in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=row[0].row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0000"
    
    def _manage_upload_flag(self, file_path, action="read", value=None):
        
        try:
            flag_path = file_path + ".upload_flag"
            if action == "read":
                if os.path.exists(flag_path):
                    with open(flag_path, 'r') as f:
                        return f.read().strip() == "True"
                return False
                
            elif action == "write":
                if value is None:
                    raise ValueError()
                with open(flag_path, 'w') as f:
                    f.write(str(value))
                return None
                
            elif action == "clear":
                if os.path.exists(flag_path):
                    os.remove(flag_path)
                return None
                
            else:
                raise ValueError(f"Not a valid action.")
                
        except Exception as e:
            if action == "read":
                return False
            return None

    def apply_formulas(self, file_path, formulas_json_path, country, models, fuels, expected_sheets):

        just_uploaded = self._manage_upload_flag(file_path, "read")
        
        try:            
            with open(formulas_json_path) as f:
                formulas_json = json.load(f)

            expanded_main_json = self._expand_main_keys_json(formulas_json, country, models, fuels, expected_sheets)

            file_path_norm = os.path.normpath(file_path)
            if file_path_norm not in expanded_main_json:
                return
            
            formulas_sheet_file = expanded_main_json[file_path_norm]
            
            specific_values = self._extract_specific_values(file_path_norm, models, fuels, expected_sheets)
            
            expanded_formulas = self._expand_single_formulas(
                formulas_sheet_file, country, models, fuels, expected_sheets, specific_values, just_uploaded
            )
            
            year_range = self._get_config_value(country, "COUNTRY_YEAR_RANGES")
            self._process_formulas_sheet_file(file_path_norm, expanded_formulas, year_range)
                
            if just_uploaded:
                self._manage_upload_flag(file_path, "clear")
                    
        except Exception as e:
            traceback.print_exc()
            raise RuntimeError(f"Error applying formulas: {str(e)}")
    
    def _expand_main_keys_json(self, formulas_json, country, models, fuels, expected_sheets):    
        
        expanded_json = {}
        for file_raw_key, file_formulas in formulas_json.items():
            
            needs_country = "{country}" in file_raw_key
            needs_model = "{model}" in file_raw_key
            needs_fuel = "{fuel}" in file_raw_key
            needs_sheet = "{sheet}" in file_raw_key
            
            combinations = product(
                [country] if needs_country else [None],
                models if needs_model else [None],
                fuels if needs_fuel else [None],
                expected_sheets if needs_sheet else [None]
            )
            
            for country_val, model_val, fuel_val, sheet_val in combinations:
                expanded_key = file_raw_key
                if country_val: expanded_key = expanded_key.replace("{country}", country_val)
                if model_val: expanded_key = expanded_key.replace("{model}", model_val)
                if fuel_val: expanded_key = expanded_key.replace("{fuel}", fuel_val)
                if sheet_val: expanded_key = expanded_key.replace("{sheet}", sheet_val)
                
                expanded_json[expanded_key] = file_formulas
                
        return expanded_json
    
    def _expand_single_formulas(self, formulas_sheet_file, country, models, fuels, expected_sheets, specific_values=None, just_uploaded=False):

        expanded_formulas = {}
        if specific_values is None:
            specific_values = {'model': None, 'fuel': None, 'sheet': None}
        
        for sheet_name, formulas in formulas_sheet_file.items():
            
            if sheet_name == "{fuel}":
                sheets_to_expand = fuels
            elif sheet_name == "{sheet}":
                sheets_to_expand = expected_sheets
            else:
                sheets_to_expand = [sheet_name]
                
            for expanded_sheet_name in sheets_to_expand:
                if expanded_sheet_name not in expanded_formulas:
                    expanded_formulas[expanded_sheet_name] = []
                
                for formula in formulas:
                    original_target = formula.get("target", "")
                    original_sources = formula.get("sources", [])

                    uses_upload_file = any("upload-" in source for source in original_sources)
                    if uses_upload_file and not just_uploaded:
                        continue
                    
                    needs_model = any("{model}" in s for s in [original_target] + original_sources)
                    if needs_model:
                        if specific_values['model']:
                            models_to_use = [specific_values['model']]
                        else:
                            models_to_use = models
                    else:
                        models_to_use = [None]
                    
                    for model_val in models_to_use:
                        new_formula = deepcopy(formula)
                        target = new_formula.get("target", "")
                        sources = new_formula.get("sources", [])
                        
                        if "{country}" in target:
                            new_formula["target"] = target.replace("{country}", country)
                        
                        if "{model}" in target and model_val:
                            new_formula["target"] = target.replace("{model}", model_val)

                        fuel_to_use = expanded_sheet_name
                        for fuel in fuels:
                            if fuel in expanded_sheet_name:
                                fuel_to_use = fuel
                         
                        if "{fuel}" in target:
                            new_formula["target"] = target.replace("{fuel}", fuel_to_use)
                        
                        if "{sheet}" in target:
                            new_formula["target"] = target.replace("{sheet}", expanded_sheet_name)
                        
                        new_sources = []
                        for source in sources:
                            if "{country}" in source:
                                source = source.replace("{country}", country)
                            if "{model}" in source and model_val:
                                source = source.replace("{model}", model_val)
                            if "{fuel}" in source:
                                source = source.replace("{fuel}", fuel_to_use)
                            if "{sheet}" in source:
                                source = source.replace("{sheet}", expanded_sheet_name)
                                
                            new_sources.append(source)

                        new_formula["sources"] = new_sources
                        expanded_formulas[expanded_sheet_name].append(new_formula)
        
        return expanded_formulas
    
    def _extract_specific_values(self, file_path, models, fuels, expected_sheets):
        
        specific_values = {
            'model': None,
            'fuel': None, 
            'sheet': None
        }
        
        for model in models:
            if model in file_path:
                specific_values['model'] = model
                break
        
        for fuel in fuels:
            if fuel in file_path:
                specific_values['fuel'] = fuel
                break
        
        for sheet in expected_sheets:
            if sheet in file_path:
                specific_values['sheet'] = sheet
                break
        
        return specific_values
    
    def _process_formulas_sheet_file(self, file_path, formulas_sheet_file, year_range):
        
        self.clear_workbook_cache()
        wb = load_workbook(file_path)
        
        try:
            changes_made = False
            for sheet_name, formulas in formulas_sheet_file.items():

                self.previously_calculated = {}
                
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]                
                for formula in formulas:
                    try:
                        formula_changed = self._execute_single_formula(ws, formula, year_range, file_path)
                        if formula_changed:
                            changes_made = True
                    except Exception as e:
                        continue
            
            if changes_made:
                #self._apply_number_formatting(ws, year_range)
                wb.save(file_path)
            
        except Exception as e:
            traceback.print_exc()
        finally:
            if hasattr(wb, 'close'):
                wb.close()
    
    def _execute_single_formula(self, ws, formula, year_range, file_path):
        
        try:
            target_label = formula["target"]
            source_labels = formula["sources"]
            formula_steps = formula.get("formula_steps", [])
            
            target_cells = self._find_cells_from_target_label(target_label, ws, year_range, file_path)
            if not target_cells:
                return False

            source_cells_list = []
            for source_label in source_labels:
                cell_refs = self._get_cell_refs_from_source_label(source_label, ws, year_range)
                if not cell_refs:
                    if "config.json" in source_label:
                        source_cells_list.append([source_label])
                    else:
                        return False
                source_cells_list.append(cell_refs)

            if target_label not in self.previously_calculated:
                self.previously_calculated[target_label] = [None] * len(target_cells)
            
            changes_made = False
            for cell_index in range(len(target_cells)):
                current_values = []
                for cell_refs in source_cells_list:
                    if cell_refs and len(cell_refs) > 0:
                        if isinstance(cell_refs[0], str) and cell_refs[0].startswith("CACHED::"):
                            label = cell_refs[0].split("::")[-1]
                            if label in self.previously_calculated:
                                if cell_index < len(self.previously_calculated[label]):
                                    value = self.previously_calculated[label][cell_index]
                                else:
                                    value = 0
                            else:
                                value = 0
                        else:
                            if "config.json" in str(cell_refs[0]):
                                ref = cell_refs[0]
                                value = self._get_cell_value_from_ref(ref, ws, cell_index)
                            else:
                                if cell_index < len(cell_refs):
                                    ref = cell_refs[cell_index]
                                    value = self._get_cell_value_from_ref(ref, ws, cell_index)
                                else:
                                    value = 0
                    else:
                        value = 0
                    current_values.append(value)
                
                final_value = self._execute_formula_for_cell(
                    formula_steps, current_values, source_cells_list, cell_index, ws
                )
                
                self.previously_calculated[target_label][cell_index] = final_value
                target_cell = target_cells[cell_index]
                current_target_value = ws[target_cell].value
                
                if current_target_value != final_value:
                    ws[target_cell] = final_value
                    changes_made = True
                    if isinstance(final_value, str):
                        ws[target_cell].number_format = '@'

            return changes_made

        except Exception as e:
            traceback.print_exc()
            return False  
    
    def _get_cell_value_from_ref(self, cell_ref, default_ws, cell_index=None):
        try:
            if "config.json" in cell_ref:
                parts = cell_ref.split("::")
                if len(parts) == 3:
                    _, country, config_key = parts
                    return self._get_config_value(country, config_key)
                elif len(parts) == 5:
                    _, country, config_key, model, fuel = parts
                    return self._get_config_value(country, config_key, model, fuel, cell_index)
                else:
                    return 0
                
            elif "::" in cell_ref and "!" in cell_ref:
                file_part, rest = cell_ref.split("::")
                sheet_part, coord = rest.split("!")
                file_path = os.path.normpath(file_part)
                
                if not os.path.exists(file_path):
                    return 0
                
                try:
                    wb = self._get_workbook(file_path, data_only=True)
                    
                    if sheet_part not in wb.sheetnames:
                        return 0
                    
                    ws = wb[sheet_part]
                    value = ws[coord].value
                    
                    cell_row = ws[coord].row
                    cell_col = ws[coord].column
                    
                    row_vals = []
                    for col in range(max(1, cell_col-2), min(ws.max_column, cell_col+2)):
                        row_vals.append(f"{ws.cell(row=cell_row, column=col).coordinate}='{ws.cell(row=cell_row, column=col).value}'")
                    
                except Exception as e:
                    return 0
            else:
                value = default_ws[cell_ref].value
            
            result = self._convert_to_numeric(value) if self._is_numeric_value(value) else 0

            return result
            
        except Exception as e:
            return 0
    
    def _find_year_column(self, ws, year_range):
        
        try:
            start_year = year_range.get('start', 0)
            if not start_year:
                return None
            
            next_year = start_year + 1
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    
                    if cell_value is None:
                        continue
                    
                    is_start_year = False
                    if str(cell_value) == str(start_year):
                        is_start_year = True
                    elif isinstance(cell_value, (int, float)) and int(cell_value) == start_year:
                        is_start_year = True
                    
                    if is_start_year:
                        if col + 1 <= ws.max_column:
                            next_cell = ws.cell(row=row, column=col + 1).value
                            if next_cell is not None:
                                if str(next_cell) == str(next_year):
                                    return col
                                elif isinstance(next_cell, (int, float)) and int(next_cell) == next_year:
                                    return col
            
            return None
            
        except Exception:
            return None
    
    def _get_cell_refs_from_source_label(self, source_label, default_ws, year_range):
        
        if "::" in source_label:
            parts = source_label.split("::")
            if len(parts) == 3:
                file_part, sheet_part, label_part = parts
            else:
                if "config.json" in source_label:
                    return [source_label]
                return []
        else:
            file_part = None
            sheet_part = default_ws.title
            label_part = source_label
        
        if label_part in self.previously_calculated:
            return [f"CACHED::{label_part}"]
        
        if "config.json" in source_label:
            return [source_label]
        
        if file_part:
            file_path = os.path.normpath(file_part)
            try:
                wb = self._get_workbook(file_path, data_only=True)
            except Exception as e:
                return []
        else:
            wb = self._get_workbook(default_ws.parent.path, data_only=True)
        
        if sheet_part not in wb.sheetnames:
            return []
        
        ws = wb[sheet_part]
        label_parts = label_part.split(":")
        
        target_row = None
        for row in range(1, ws.max_row + 1):
            match = True
            for col_idx, expected_value in enumerate(label_parts, start=1):
                cell_value = ws.cell(row=row, column=col_idx).value
                cell_str = str(cell_value).strip() if cell_value is not None else ""
                expected_str = expected_value.strip()
                
                if "vs. Baseline" not in expected_str:
                    if expected_str != cell_str:
                        match = False
                        break
                else:
                    if expected_str not in cell_str:
                        match = False
                        break
            
            if match:
                target_row = row
                break
        
        if target_row is None:
            return []
        
        if not year_range:
            return []
        
        start_col = self._find_year_column(ws, year_range,)
        if not start_col:
            return []
        
        cell_refs = []
        for col in range(start_col, ws.max_column + 1):
            cell = ws.cell(row=target_row, column=col)
            cell_value = cell.value
            
            if cell_value is None:
                break
                
            if file_part:
                ref = f"{file_path}::{sheet_part}!{cell.coordinate}"
            else:
                ref = cell.coordinate
                
            cell_refs.append(ref)
        
        return cell_refs
    
    def _find_cells_from_target_label(self, target_label, worksheet, year_range, file_path):
        
        label_parts = target_label.split(":")
        target_row = None

        for row in range(1, min(21, worksheet.max_row + 1)):
            type_value = str(worksheet.cell(row=row, column=1).value or "").strip()
            sub_value = str(worksheet.cell(row=row, column=2).value or "").strip()
            
        for row in range(1, worksheet.max_row + 1):
            type_value = str(worksheet.cell(row=row, column=1).value or "").strip()
            sub_value = str(worksheet.cell(row=row, column=2).value or "").strip()
            expected_type = label_parts[0].strip()
            expected_sub = label_parts[1].strip() if len(label_parts) > 1 else ""
            
            is_vs_baseline_label = "vs. Baseline" in target_label
            if not is_vs_baseline_label:
                type_match = (expected_type == type_value)
            else:
                type_match = (expected_type in type_value)
            
            sub_match = (expected_sub == "") or (expected_sub == sub_value)
            
            if type_match and sub_match:
                target_row = row
                break

        if target_row is None:
            return []
        
        row_vals = []
        for col in range(1, min(10, worksheet.max_column + 1)):
            cell = worksheet.cell(row=target_row, column=col)
            row_vals.append(f"{cell.coordinate}='{cell.value}'")
        
        start_col = self._find_year_column(worksheet, year_range)

        is_amount_target = False
        if "design" in file_path:
            for amount_target in self.amount_targets:
                if amount_target == target_label:
                    is_amount_target = True
                    break
        
        if is_amount_target:
            return [f"C{target_row}"]
        
        if start_col is None:
            start_search_col = len(label_parts) + 1
            for col in range(start_search_col, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=target_row, column=col).value
                if self._is_numeric_value(cell_value):
                    start_col = col
                    break
        
        if start_col is None:
            return []
        
        numeric_cells = []        
        for col in range(start_col, worksheet.max_column + 1):
            cell = worksheet.cell(row=target_row, column=col)
            cell_value = cell.value
            numeric_cells.append(cell.coordinate)

        if year_range:
            expected_years = year_range['end'] - year_range['start'] + 1
            
            if len(numeric_cells) > expected_years:                
                numeric_cells = numeric_cells[:expected_years]
                
        return numeric_cells
    
    def _get_config_value(self, country, config_key, model=None, fuel=None, cell_index=None):
        try:
            if not os.path.exists(CONFIG_FILE):
                return 0
            
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
            
            if ":" in config_key:
                main_key, sub_key = config_key.split(":", 1)
            else:
                main_key, sub_key = config_key, None
            
            if main_key not in config:
                return 0
            
            main_config = config[main_key]
            
            if model and fuel:
                if country in main_config:
                    country_data = main_config[country]
                    
                    if model in country_data:
                        model_data = country_data[model]
                        if isinstance(model_data, dict) and fuel in model_data:
                            fuel_data = model_data[fuel]
                            
                            if sub_key:
                                value = fuel_data.get(sub_key)
                                if value is not None:
                                    return float(value)
                            else:
                                if cell_index is not None and isinstance(fuel_data, dict):
                                    years = list(fuel_data.keys())
                                    if years and 0 <= cell_index < len(years):
                                        year = years[cell_index]
                                        value = fuel_data.get(year)
                                        if value is not None:
                                            return float(value)
                                
                                if isinstance(fuel_data, dict):
                                    return fuel_data
                                else:
                                    return float(fuel_data)
                        return None
                    return None
                return 0
            
            if sub_key:
                if country in main_config:
                    country_data = main_config[country]
                    if isinstance(country_data, dict):
                        value = country_data.get(sub_key)
                        if value is not None:
                            return float(value)
            else:
                value = main_config.get(country)
                if isinstance(value, dict):
                    return value
                elif value is not None:
                    return float(value)
            
            return 0
                            
        except Exception as e:
            return 0
    
    def _execute_formula_for_cell(self, formula_steps, current_values, source_cells_list, cell_index, default_ws):
    
        results = {}

        for step in formula_steps:
            op = step["op"]
            operands = step["operands"]
            result_key = step["result"]

            if op == "get_cell":
                result = self._execute_get_cell_operation(operands, source_cells_list, default_ws)
            
            elif op == "offset":
                result = self._execute_offset_operation(
                    operands, current_values, results, source_cells_list, cell_index, default_ws
                )
            elif op == "sum_range":
                result = self._execute_sum_range_operation(
                    operands, current_values, results, source_cells_list, cell_index, default_ws, step
                )
            elif op == "set_values":
                index = self._get_operand_value(operands[0], current_values, results)
                index_start = self._get_operand_value(operands[1], current_values, results)
                value = self._get_operand_value(operands[2], current_values, results)
                if cell_index < index_start: 
                    result = index
                else:
                    result = value

            elif op == "range":
                range_value = self._get_operand_value(operands[0], current_values, results)
                result = cell_index + range_value
                    
            elif op == "value":
                result = self._execute_value_operation(
                    operands, current_values, results, source_cells_list, default_ws
                )
            
            elif op == "index_match":
                result = self._execute_index_match_operation(
                    operands, current_values, results, source_cells_list
                )
            
            else:
                ops_args = []
                for operand in operands:
                    ops_args.append(self._get_operand_value(operand, current_values, results))
                
                if op in OPS_MAP:
                    result = OPS_MAP[op](*ops_args)
                else:
                    result = 0
            
            results[result_key] = result
        
        final_result = results.get("final", 0)
        return final_result
    
    def _meets_condition(self, value, condition):
        
        if condition is None:
            return True
        
        try:
            if condition == "<0":
                return value < 0
            elif condition == ">0":
                return value > 0
            elif condition == "<=0":
                return value <= 0
            elif condition == ">=0":
                return value >= 0
            elif condition == "==0":
                return value == 0
            elif condition == "!=0":
                return value != 0
            else:
                return True
        except:
            return True

    def _get_operand_value(self, operand, current_values, current_results):
        
        try:
            if isinstance(operand, list):
                if operand[0] == "index":
                    key = operand[1]
                    if isinstance(key, int):
                        value = current_values[key] if key < len(current_values) else 0
                    else:
                        value = current_results.get(key, 0)
                    return float(value) if value is not None else 0
                elif operand[0] in ["literal", "range"]:
                    value = operand[1]
                    return float(value) if value is not None else 0
                elif operand[0] == "str":
                    value = operand[1]
                    return str(value) if value is not None else ""
            elif isinstance(operand, str):
                value = current_results.get(operand, 0)
                return float(value) if value is not None else 0
            elif isinstance(operand, (int, float)):
                return float(operand)
            else:
                return 0
        except (ValueError, TypeError) as e:
            return 0
    
    def _execute_get_cell_operation(self, operands, source_cells_list, default_ws):

        if operands[0][0] == "index" and operands[1][0] == "literal":
            source_index = int(operands[0][1])
            literal_index = int(operands[1][1])
        else:
            return 0

        refs = source_cells_list[source_index]
        first_ref = refs[0]

        if isinstance(first_ref, str) and first_ref.startswith("CACHED::"):
            label = first_ref.split("::")[-1]
            if label in self.previously_calculated:
                return self.previously_calculated[label][0]

        if "!" in first_ref:
            sheet_part, cell_ref = first_ref.split("!", 1)

        col_letter, row_num = coordinate_from_string(cell_ref)
        start_col = column_index_from_string(col_letter)
        
        result = []
        for c in range(1, start_col):
            result.append(f"{sheet_part}!{get_column_letter(c)}{row_num}")

        result.extend(refs)
        cell_ref = result[literal_index]
        cell_value = self._get_cell_value_from_ref(cell_ref, default_ws)

        return cell_value

    def _execute_offset_operation(self, operands, current_values, current_results, source_cells_list, current_index, default_ws):

        if isinstance(operands[0], list) and operands[0][0] == "index":
            source_index = operands[0][1]
        else:
            source_index = self._get_operand_value(operands[0], current_values, current_results)

        if isinstance(operands[1], list) and operands[1][0] == "literal":
            offset = operands[1][1]
        else:
            offset = self._get_operand_value(operands[1], current_values, current_results)

        source_index = int(source_index)
        offset = int(offset)

        if not (0 <= source_index < len(source_cells_list)):
            return 0

        cell_refs = source_cells_list[source_index]
        target_index = current_index - offset

        if target_index < 0:
            return 0

        if cell_refs[0].startswith("CACHED::"):
            label = cell_refs[0].split("::")[-1]

            if label in self.previously_calculated:
                cached_values = self.previously_calculated[label]
                return cached_values[target_index] if target_index < len(cached_values) else 0
            else:
                return 0

        if target_index < len(cell_refs):
            return self._get_cell_value_from_ref(cell_refs[target_index], default_ws)

        return 0
    
    def _execute_sum_range_operation(self, operands, current_values, current_results, source_cells_list, current_index, default_ws, step=None):
    
        if isinstance(operands[0], list) and operands[0][0] == "index":
            source_index = operands[0][1]
        else:
            source_index = self._get_operand_value(operands[0], current_values, current_results)

        if isinstance(operands[1], list) and operands[1][0] == "literal":
            width = operands[1][1]
        else:
            width = self._get_operand_value(operands[1], current_values, current_results)

        direction = step.get("direction", "forward") if step else "forward"
        condition = step.get("condition") if step else None
        
        try:
            source_index = int(source_index)
        except (TypeError, ValueError) as e:
            return 0
        
        try:
            current_index = int(current_index)
        except (TypeError, ValueError) as e:
            return 0
        
        if width != "all":
            try:
                width = int(width)
            except (TypeError, ValueError) as e:
                return 0
        
        if not (0 <= source_index < len(source_cells_list)):
            return 0

        cell_refs = source_cells_list[source_index]
        
        if not cell_refs:
            return 0

        actual_values = []
        if cell_refs[0].startswith("CACHED::"):
            label = cell_refs[0].split("::")[-1]
            actual_values = self.previously_calculated.get(label, [])
        else:
            for i, ref in enumerate(cell_refs):
                value = self._get_cell_value_from_ref(ref, default_ws)
                actual_values.append(value)
                
        def get_value(i):
            if 0 <= i < len(actual_values):
                val = actual_values[i]
                return val
            return 0
        
        if width == "all":
            total = 0
            for i in range(len(actual_values)):
                value = get_value(i)
                meets = self._meets_condition(value, condition)
                if meets:
                    total += value
            return total

        total = 0
        
        if direction == "backward":
            if width == 0:
                start_index = 0
                end_index = current_index + 1
            else:
                start_index = max(0, current_index - width + 1)
                end_index = current_index + 1
            
            start_index = int(start_index)
            end_index = int(end_index)
            
            for i in range(start_index, end_index):
                value = get_value(i)
                meets = self._meets_condition(value, condition)
                if meets:
                    total += value
                    
        elif direction == "forward":
            
            from_start = step.get("from_start", False) if step else False
            
            if from_start:
                start_index = 0
                end_index = min(len(actual_values), width)
            else:
                start_index = current_index
                end_index = min(len(actual_values), current_index + width)
            
            start_index = int(start_index)
            end_index = int(end_index)
            
            total = 0
            for i in range(start_index, end_index):
                value = get_value(i)
                if self._meets_condition(value, condition):
                    total += value
            
            return total
        
        return total
    
    def _execute_value_operation(self, operands, current_values, current_results, source_cells_list, default_ws):

        if len(operands) >= 2:

            if isinstance(operands[0], list) and operands[0][0] == "index":
                source_index = operands[0][1]
            else:
                source_index = self._get_operand_value(operands[0], current_values, current_results)

            if isinstance(operands[1], list) and operands[1][0] == "literal":
                cell_index_to_get = operands[1][1]
            else:
                cell_index_to_get = self._get_operand_value(operands[1], current_values, current_results)

            source_index = int(source_index)
            cell_index_to_get = int(cell_index_to_get)

            if not (0 <= source_index < len(source_cells_list)):
                return 0

            cell_refs = source_cells_list[source_index]
            
            if cell_refs[0].startswith("CACHED::"):
                label = cell_refs[0].split("::")[-1]
                return self.previously_calculated[label][cell_index_to_get] if cell_index_to_get < len(self.previously_calculated[label]) else 0
                    
            if 0 <= cell_index_to_get < len(cell_refs):
                return self._get_cell_value_from_ref(cell_refs[cell_index_to_get], default_ws)

        return 0
    
    def _execute_index_match_operation(self, operands, current_values, current_results,
     source_cells_list):
        
        if isinstance(operands[0], list) and operands[0][0] == "index":
            source_index = operands[0][1]
        else:
            source_index = self._get_operand_value(operands[0], current_values, current_results)

        cell_refs = source_cells_list[source_index]
        if cell_refs[0].startswith("CACHED::"):
            label = cell_refs[0].split("::")[-1]
            if label in self.previously_calculated:
                cached_values = self.previously_calculated[label]
                for i, value in enumerate(cached_values):
                    if value != 0:
                        return i

        return 0
    