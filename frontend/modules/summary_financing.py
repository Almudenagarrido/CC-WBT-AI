import os
import json
import streamlit as st
from itertools import product
import plotly.graph_objects as go
from copy import deepcopy
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.dirname(BASE_DIR)
PROJECT_ROOT = os.path.dirname(FRONTEND_DIR)
BACKEND_DIR = os.path.join(PROJECT_ROOT, "backend")
CONFIG_FILE = os.path.join(BACKEND_DIR, "config.json")
VISUALIZATIONS_JSON = os.path.join(BASE_DIR, "visualizations_map.json")

OPS_MAP = {
    "copy": lambda x: x,
    "addition": lambda x, y: x + y,
    "negative": lambda x: -x,
}


class SummaryFinancing:
    
    def __init__(self, country):
        self.country = country
        self.formulas_json_path = VISUALIZATIONS_JSON
        self.config = self._load_config()
        self.year_range = self._get_year_range()
        self.previously_calculated = {}
        self._workbook_cache = {}
        self.graph_filters = {"Sources of Financing": ["Baseline"], "Changes in the Capital Structure": ["Baseline"]}
        self.value_filters = {"Revenues -": ["Baseline"], "EBITDA -": ["Baseline"], "Grants to CAPEX -": ["Baseline"], "New Debt -": ["Baseline"], "Equity -": ["Baseline"], "CAPEX -": ["Baseline"], "Potential Income from Carbon Credits": ["Baseline"]}
        self.line_values = ["Equity -", "Potential Income from Carbon Credits"]
        self.color_palette = [
            "#14027D",
            "#7b018e",
            "#fca818",
            "#06aa91",
            "#f1e255",
        ]
    
    def _load_config(self):
        with open(CONFIG_FILE, "r", encoding='utf-8') as f:
            return json.load(f)
    
    def _get_year_range(self):
        year_ranges = self.config.get("COUNTRY_YEAR_RANGES", {})
        country_range = year_ranges.get(self.country, {})
        return {
            "start": country_range.get("start", 2020),
            "end": country_range.get("end", 2050)
        }
    
    def _get_workbook(self, file_path, data_only=True):
        if file_path not in self._workbook_cache:
            try:
                self._workbook_cache[file_path] = load_workbook(file_path, data_only=data_only)
            except Exception:
                return None
        return self._workbook_cache[file_path]
    
    def _is_numeric_value(self, value):
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value.replace(",", ".").strip())
                return True
            except ValueError:
                return False
        return False
    
    def _convert_to_numeric(self, value):
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", ".").strip())
            except ValueError:
                return 0
        return 0
    
    def _expand_visualizations_json(self, visualizations_json):
        
        expanded_graphs = {}
        models = self.config.get("MODELS", {}).get(self.country, [])
        fuels = self.config.get("FUELS", {}).get(self.country, {}).get("normal", [])
        sheets = self.config.get("FUELS", {}).get(self.country, {}).get("expanded", [])
        more_sheets = self.config.get("FUELS", {}).get(self.country, {}).get("more_expanded", [])
        
        for graph_raw_name, graph_data in visualizations_json.items():

            has_country = "{country}" in graph_raw_name
            has_model = "{model}" in graph_raw_name
            has_fuel = "{fuel}" in graph_raw_name
            has_sheet = "{sheet}" in graph_raw_name
            has_more_sheet = "{more_sheet}" in graph_raw_name
            
            if not (has_country or has_model or has_fuel or has_sheet or has_more_sheet):
                expanded_graphs[graph_raw_name] = deepcopy(graph_data)
                continue
            
            country_values = [self.country]
            model_values = models if has_model else [""]
            fuel_values = fuels if has_fuel else [""]
            sheet_values = sheets if has_sheet else [""]
            more_sheet_values = more_sheets if has_more_sheet else [""]

            for country_val, model_val, fuel_val, sheet_val, more_sheet_val in product(country_values, model_values, fuel_values, sheet_values, more_sheet_values):
                expanded_name = graph_raw_name
                
                if not fuel_val and "Sources of Financing" in expanded_name:
                    fuel_val = "Electricity"
                if has_sheet and fuel_val not in sheet_val:
                    continue
                if has_more_sheet and fuel_val not in more_sheet_val:
                    continue

                if has_country:
                    expanded_name = expanded_name.replace("{country}", country_val)
                if has_model:
                    expanded_name = expanded_name.replace("{model}", model_val)
                if has_fuel:
                    expanded_name = expanded_name.replace("{fuel}", fuel_val)
                if has_sheet:
                    expanded_name = expanded_name.replace("{sheet}", sheet_val)
                if has_more_sheet:
                    expanded_name = expanded_name.replace("{more_sheet}", more_sheet_val)
                
                if expanded_name != graph_raw_name:
                    
                    if expanded_name in expanded_graphs:
                        
                        existing_graph = expanded_graphs[expanded_name]
                        new_sources = graph_data.get("financing_sources", {})
                        
                        for source_name, source_info in new_sources.items():
                            existing_graph["financing_sources"][source_name] = deepcopy(source_info)
                        
                        continue
                    
                    expanded_graph_data = deepcopy(graph_data)
                    expansion_info = {
                        'country': country_val,
                        'model': model_val,
                        'fuel': fuel_val,
                        'sheet': sheet_val,
                        'more_sheet': more_sheet_val
                    }
                    expanded_graph_data['_expansion_info'] = expansion_info
                    expanded_graphs[expanded_name] = expanded_graph_data
        
        return expanded_graphs

    def _expand_sources_paths(self, expanded_graphs):
        
        models = self.config.get("MODELS", {}).get(self.country, [])
        
        for _, graph_data in expanded_graphs.items():
            
            if "financing_sources" not in graph_data:
                continue
            
            financing_sources = graph_data["financing_sources"]
            expansion_info = graph_data.get('_expansion_info', {})
            country_val = expansion_info.get('country', self.country)
            model_val = expansion_info.get('model', '')
            fuel_val = expansion_info.get('fuel', '')
            sheet_val = expansion_info.get('sheet', '')
            more_sheet_val = expansion_info.get('more_sheet', '')
            
            model_values = [model_val] if model_val else models
            
            all_expanded_sources = {}
            
            for current_model_val in model_values:
                
                for financing_source, finance_source_info in financing_sources.items():
                    
                    expanded_source_name = financing_source
                    if "{model}" in expanded_source_name:
                        expanded_source_name = expanded_source_name.replace("{model}", current_model_val)
                    
                    if "{country}" in expanded_source_name:
                        expanded_source_name = expanded_source_name.replace("{country}", country_val)
                    if "{fuel}" in expanded_source_name:
                        expanded_source_name = expanded_source_name.replace("{fuel}", fuel_val)
                    if "{sheet}" in expanded_source_name:
                        expanded_source_name = expanded_source_name.replace("{sheet}", sheet_val)
                    if "{more_sheet}" in expanded_source_name:
                        expanded_source_name = expanded_source_name.replace("{more_sheet}", more_sheet_val)
                    
                    if expanded_source_name in all_expanded_sources:
                        continue
                    
                    if "data_sources" not in finance_source_info:
                        all_expanded_sources[expanded_source_name] = finance_source_info
                        continue
                    
                    data_sources_paths = finance_source_info["data_sources"]
                    expanded_paths = []
                    
                    for data_source_path in data_sources_paths:
                        expanded_path = data_source_path

                        if "{country}" in expanded_path:
                            expanded_path = expanded_path.replace("{country}", country_val)
                        if "{model}" in expanded_path:
                            expanded_path = expanded_path.replace("{model}", current_model_val)
                        if "{fuel}" in expanded_path:
                            expanded_path = expanded_path.replace("{fuel}", fuel_val)
                        if "{sheet}" in expanded_path:
                            expanded_path = expanded_path.replace("{sheet}", sheet_val)
                        if "{more_sheet}" in expanded_path:
                            expanded_path = expanded_path.replace("{more_sheet}", more_sheet_val)
                        
                        parts = expanded_path.split("::", 1)
                        file_path_part = parts[0]
                        metadata_part = parts[1]
                        
                        if "\\" in file_path_part:
                            parts_file = file_path_part.split("\\", 1)
                            if len(parts_file) == 2:
                                country_dir, rest_file_path = parts_file
                                full_file_path = os.path.join(BACKEND_DIR, country_dir, rest_file_path)
                            else:
                                full_file_path = os.path.join(BACKEND_DIR, country_val, file_path_part.lstrip("\\"))
                        else:
                            full_file_path = os.path.join(BACKEND_DIR, country_val, file_path_part)
                        
                        full_file_path = os.path.normpath(full_file_path)
                        full_path = f"{full_file_path}::{metadata_part}"
                        
                        expanded_paths.append(full_path)
                    
                    finance_source_info_copy = dict(finance_source_info)
                    finance_source_info_copy["data_sources"] = expanded_paths
                    
                    all_expanded_sources[expanded_source_name] = finance_source_info_copy
            
            graph_data["financing_sources"] = all_expanded_sources
                    
        return expanded_graphs

    def _find_year_column(self, ws, year_range):
        try:
            start_year = year_range.get('start', 0)
            if not start_year:
                return None
            
            next_year = start_year + 1
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    
                    if cell_value is None:
                        continue
                    
                    is_start_year = False
                    if str(cell_value) == str(start_year):
                        is_start_year = True
                    elif isinstance(cell_value, (int, float)) and int(cell_value) == start_year:
                        is_start_year = True
                    
                    if is_start_year:
                        if col + 1 <= ws.max_column:
                            next_cell = ws.cell(row=row, column=col + 1).value
                            if next_cell is not None:
                                if str(next_cell) == str(next_year):
                                    return col
                                elif isinstance(next_cell, (int, float)) and int(next_cell) == next_year:
                                    return col
            
            return None
        except Exception:
            return None
    
    def _get_cell_refs_from_source_label(self, source_path, year_range):
        try:
            if "::" not in source_path:
                return []
            
            file_part, sheet_part, label_part = source_path.split("::", 2)
            
            if not os.path.exists(file_part):
                return []
            
            wb = self._get_workbook(file_part)
            if wb is None:
                return []
            
            if sheet_part not in wb.sheetnames:
                return []
            
            ws = wb[sheet_part]
            label_parts = label_part.split(":")
            
            target_row = None
            for row in range(1, ws.max_row + 1):
                match = True
                
                for col_idx, expected_value in enumerate(label_parts, start=1):
                    cell_value = ws.cell(row=row, column=col_idx).value
                    cell_str = str(cell_value).strip() if cell_value is not None else ""
                    expected_str = expected_value.strip()
                    
                    if "vs. Baseline" not in expected_str:
                        if cell_str != expected_str:
                            match = False
                            break
                    else:
                        if expected_str not in cell_str:
                            match = False
                            break
                
                if match:
                    target_row = row
                    break
            
            if target_row is None:
                return []
            
            start_col = self._find_year_column(ws, year_range)
            if not start_col:
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=1, column=col).value
                    if val == year_range['start']:
                        start_col = col
                        break
            
            if not start_col:
                return []
            
            num_years = year_range['end'] - year_range['start'] + 1
            cell_refs = []
            
            for offset in range(num_years):
                col = start_col + offset
                if col > ws.max_column:
                    break
                
                cell_coord = ws.cell(row=target_row, column=col).coordinate
                cell_refs.append(f"{file_part}::{sheet_part}!{cell_coord}")
            
            return cell_refs
            
        except Exception:
            return []
    
    def _get_cell_value_from_ref(self, cell_ref):
        try:
            if "::" in cell_ref and "!" in cell_ref:
                file_part, rest = cell_ref.split("::", 1)
                sheet_part, coord = rest.split("!", 1)
                
                if not os.path.exists(file_part):
                    return 0
                
                wb = self._get_workbook(file_part)
                if wb is None or sheet_part not in wb.sheetnames:
                    return 0
                
                ws = wb[sheet_part]
                value = ws[coord].value
                
                if self._is_numeric_value(value):
                    num_value = self._convert_to_numeric(value)
                    return num_value
                else:
                    return 0
            else:
                return 0
                
        except Exception:
            return 0
    
    def _get_operand_value(self, operand, current_values, current_results):
        try:
            if isinstance(operand, list):
                if operand[0] == "index":
                    key = operand[1]
                    if isinstance(key, int):
                        value = current_values[key] if key < len(current_values) else 0
                    else:
                        value = current_results.get(key, 0)
                    return float(value) if value is not None else 0
                elif operand[0] in ["literal", "range"]:
                    value = operand[1]
                    return float(value) if value is not None else 0
                elif operand[0] == "str":
                    value = operand[1]
                    return str(value) if value is not None else ""
            elif isinstance(operand, str):
                value = current_results.get(operand, 0)
                return float(value) if value is not None else 0
            elif isinstance(operand, (int, float)):
                return float(operand)
            else:
                return 0
        except (ValueError, TypeError):
            return 0
    
    def _execute_formula_steps(self, formula_steps, current_values, source_cells_list, cell_index):
        
        if not formula_steps:
            return current_values[0] if current_values else 0
        
        results = {}
        
        for step in formula_steps:
            op = step["op"]
            operands = step["operands"]
            result_key = step["result"]
            
            if op == "copy":
                ops_args = []
                for operand in operands:
                    arg_value = self._get_operand_value(operand, current_values, results)
                    ops_args.append(arg_value)
                
                result = OPS_MAP[op](*ops_args)
                results[result_key] = result
                
            elif op == "index_copy":
                
                position_operand = operands[1]
                
                position = position_operand[1]
                value = self._get_cell_value_from_ref(source_cells_list[0][position])

                if cell_index == (len(source_cells_list[0]) + position):
                    results[result_key] = value
                else:
                    results[result_key] = 0
                                    
            else:
                ops_args = []
                for operand in operands:
                    arg_value = self._get_operand_value(operand, current_values, results)
                    ops_args.append(arg_value)
                
                if op in OPS_MAP:
                    try:
                        result = OPS_MAP[op](*ops_args)
                        results[result_key] = result
                    except Exception:
                        results[result_key] = 0
                else:
                    results[result_key] = 0
        
        return results.get("final", 0)
        
    def _process_source_data(self, source_info):

        formula_steps = source_info.get("formula_steps", [])
        data_sources_paths = source_info.get("data_sources", [])

        if not data_sources_paths:
            return [0 for _ in range(self.year_range["start"], self.year_range["end"] + 1)]
        
        source_cells_list = []
        for source_path in data_sources_paths:

            cell_refs = self._get_cell_refs_from_source_label(source_path, self.year_range)
            
            if not cell_refs:
                return [0 for _ in range(self.year_range["start"], self.year_range["end"] + 1)]
            
            source_cells_list.append(cell_refs)
        
        num_years = self.year_range['end'] - self.year_range['start'] + 1
        calculated_values = []
        
        for cell_index in range(num_years):
            current_values = []
            for cell_refs in source_cells_list:
                if cell_index < len(cell_refs):
                    value = self._get_cell_value_from_ref(cell_refs[cell_index])
                    current_values.append(value)
                else:
                    current_values.append(0)

            final_value = self._execute_formula_steps(
                formula_steps, 
                current_values, 
                source_cells_list, 
                cell_index
            )
            
            calculated_values.append(final_value)
        
        return calculated_values
        
    def _calculate_graph_values(self, graph_data):
        source_values = {}
        
        if "financing_sources" in graph_data:
            financing_sources = graph_data["financing_sources"]

            for source_name, source_info in financing_sources.items():
                
                values = self._process_source_data(source_info)
                
                if values:
                    source_values[source_name] = values
        
        return source_values
    
    def _calculate_values(self):
        
        with open(self.formulas_json_path, "r", encoding='utf-8') as f:
            visualizations_json = json.load(f)
        
        expanded_graphs = self._expand_visualizations_json(visualizations_json)
        final_graphs = self._expand_sources_paths(expanded_graphs)
        final_graph_values = {}
        
        for graph_name, graph_data in final_graphs.items():

            data_source_values = self._calculate_graph_values(graph_data)

            sum_values = True if graph_data.get("sum_values") == "True" else False
            if sum_values:
                for key, values in data_source_values.items():
                    data_source_values[key] = sum(values)

            final_graph_values[graph_name] = {
                "chart_type": graph_data.get("chart_type"),
                "source_values": data_source_values,
                "years": list(range(self.year_range['start'], self.year_range['end'] + 1))
            }
        
        for wb in self._workbook_cache.values():
            if hasattr(wb, 'close'):
                wb.close()
        self._workbook_cache.clear()
        
        return final_graph_values
    
    def _render_single_value_bar_graph(self, graph_name, source_values):
    
        fig = go.Figure()
        
        source_names = list(source_values.keys())
        
        for idx, source_name in enumerate(source_names):
            values = source_values[source_name]
            
            color = self.color_palette[idx % len(self.color_palette)]
            
            fig.add_trace(go.Bar(
                name=source_name,
                x=[source_name],
                y=[values],
                text=[f"{values:,.2f}"],
                textposition='auto',
                marker_color=color
            ))
        
        fig.update_layout(
            xaxis_title="",
            yaxis_title="Total Value",
            barmode='group',
            showlegend=False,
            template="plotly_white"
        )
        
        st.plotly_chart(fig, use_container_width=True, key=f"single_bar_{graph_name}")

    def _render_yearly_bar_graph(self, graph_name, source_values, years):

        fig = go.Figure()
        
        filtered_sources = []
        for source_name in source_values.keys():
            if not self._should_skip_value(source_name):
                filtered_sources.append(source_name)
        
        for idx, source_name in enumerate(filtered_sources):
            values = source_values[source_name]
            
            if isinstance(values, list) and len(values) == len(years):
                is_line = any(line_pattern in source_name for line_pattern in self.line_values)
                
                color = self.color_palette[idx % len(self.color_palette)]
                
                if is_line:
                    fig.add_trace(go.Scatter(
                        name=source_name,
                        x=[str(year) for year in years],
                        y=values,
                        mode='lines+markers',
                        line=dict(width=3, color=color),  marker=dict(size=8, color=color),  text=[f"{v:,.2f}" for v in values],
                        hoverinfo='text+name+x'
                    ))
                else:
                    fig.add_trace(go.Bar(
                        name=source_name,
                        x=[str(year) for year in years],
                        y=values,
                        text=[f"{v:,.2f}" for v in values],
                        textposition='auto',
                        marker_color=color
                    ))
        
        fig.update_layout(
            xaxis_title="Years",
            yaxis_title="Total Value",
            barmode='group',
            showlegend=True,
            template="plotly_white",
            legend=dict(
                title="Sources",
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )
        
        st.plotly_chart(fig, use_container_width=True, key=f"yearly_bar_{graph_name}")

    def _should_skip_graph(self, graph_name):

        for pattern, exclude_list in self.graph_filters.items():
            if pattern in graph_name:
                
                for exclude_pattern in exclude_list:
                    if exclude_pattern in graph_name:
                        return True
        return False
    
    def _should_skip_value(self, source_name):

        for pattern, exclude_list in self.value_filters.items():
            if pattern in source_name:
                for exclude_pattern in exclude_list:
                    if exclude_pattern in source_name:
                        return True
        return False
    
    def __call__(self):
    
        st.write("### Summary Financing Dashboard")
        graph_values = self._calculate_values()
        
        graph_items = []
        for graph_name, graph_data in graph_values.items():
            if not self._should_skip_graph(graph_name):
                graph_items.append((graph_name, graph_data))
        
        if not graph_items:
            return
        
        total_graphs = len(graph_items)
        
        for i in range(0, total_graphs, 2):
            col1, col2 = st.columns(2, gap="large")
            
            with col1:
                
                graph_name1, graph_data1 = graph_items[i]
                st.write(f"###### {graph_name1}")
                
                source_values1 = graph_data1["source_values"]
                chart_type1 = graph_data1.get("chart_type", "bar")
                year_range1 = graph_data1.get("years", range(2020, 2061))
                
                if chart_type1 == "single_value_bar":
                    self._render_single_value_bar_graph(graph_name1, source_values1)
                elif chart_type1 == "yearly_group_bar":
                    self._render_yearly_bar_graph(graph_name1, source_values1, year_range1)

            with col2:
                if i + 1 < total_graphs:
                    
                    graph_name2, graph_data2 = graph_items[i + 1]
                    st.write(f"###### {graph_name2}")
                    
                    source_values2 = graph_data2["source_values"]
                    chart_type2 = graph_data2.get("chart_type", "bar")
                    year_range2 = graph_data2.get("years", range(2020, 2061))
                    
                    if chart_type2 == "single_value_bar":
                        self._render_single_value_bar_graph(graph_name2, source_values2)
                    elif chart_type2 == "yearly_group_bar":
                        self._render_yearly_bar_graph(graph_name2, source_values2, year_range2)
            
            if i + 2 < total_graphs:
                st.divider()  

